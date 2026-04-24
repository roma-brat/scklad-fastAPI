"""API роуты для отчётов"""
from fastapi import APIRouter, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from main import get_db, get_user
import io
import json
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/reports", tags=["reports"])
templates = Jinja2Templates(directory="templates")

# Styles for export
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
ALT_ROW_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


@router.get("/", response_class=HTMLResponse)
async def reports_page(request: Request):
    user = get_user(request)
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/dashboard", status_code=303)
    return templates.TemplateResponse("reports/list.html", {"request": request, "current_user": user})


@router.get("/export/transactions")
async def export_transactions(request: Request):
    user = get_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    db = get_db()
    try:
        transactions = db.get_transactions_dict(limit=1000)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Транзакции"
        ws.append(['Время', 'Пользователь', 'Товар', 'Тип', 'Количество', 'Детали'])
        for txn in transactions:
            ws.append([
                txn.get('timestamp',''),
                txn.get('user_name',''),
                txn.get('item_name',''),
                txn.get('operation_type',''),
                txn.get('quantity',0),
                txn.get('detail','')
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=transactions.xlsx"})
    except Exception as e:
        return HTMLResponse(content=f"Ошибка: {str(e)}", status_code=500)


@router.get("/export/full")
async def export_full_database(request: Request):
    """Экспорт всех таблиц БД в Excel (как в Flet проекте)"""
    user = get_user(request)
    if not user or user.get("role") != "admin":
        return RedirectResponse(url="/dashboard", status_code=303)

    db = get_db()
    try:
        from sqlalchemy import inspect, text
        
        # Получаем список ВСЕХ таблиц из БД
        inspector = inspect(db.engine)
        all_table_names = inspector.get_table_names()
        
        logger.info(f"Found {len(all_table_names)} tables to export")
        
        # Исключаем служебные таблицы
        EXCLUDED_TABLES = {"audit_log"}
        tables_to_export = [t for t in all_table_names if t not in EXCLUDED_TABLES]
        
        # Экспортируем каждую таблицу
        tables_data = {}
        for table_name in tables_to_export:
            try:
                with db.get_session() as session:
                    result = session.execute(text(f"SELECT * FROM {table_name}"))
                    columns = result.keys()
                    rows = result.fetchall()
                    
                    tables_data[table_name] = [
                        {col: row[idx] for idx, col in enumerate(columns)}
                        for row in rows
                    ]
                    
                    logger.info(f"Exported {table_name}: {len(rows)} rows")
            except Exception as e:
                logger.error(f"Export table '{table_name}' error: {e}")
                tables_data[table_name] = []

        # Создаём Excel файл
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for table_name, rows in tables_data.items():
            ws = wb.create_sheet(title=_sanitize_sheet_name(table_name))

            if not rows:
                ws.cell(row=1, column=1, value="Нет данных")
                ws.cell(row=1, column=1).font = Font(italic=True, color="999999")
                continue

            columns = list(rows[0].keys())

            # Header row
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGNMENT
                cell.border = THIN_BORDER

            # Data rows
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    value = row_data.get(col_name, "")
                    value = _to_excel_value(value)
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = DATA_ALIGNMENT
                    cell.border = THIN_BORDER
                    if row_idx % 2 == 0:
                        cell.fill = ALT_ROW_FILL

            # Auto column width
            for col_idx in range(1, len(columns) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

            # Auto filter and freeze panes
            ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(columns))}{len(rows) + 1}"
            ws.freeze_panes = "A2"

        # Save to memory
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
        filename = f"sklad_export_{timestamp}.xlsx"
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Exported {len(tables_data)} tables to memory")
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Full export error: {e}", exc_info=True)
        return HTMLResponse(content=f"Ошибка экспорта: {str(e)}", status_code=500)


def _sanitize_sheet_name(name: str) -> str:
    forbidden = [":", "\\", "/", "?", "*", "[", "]"]
    for ch in forbidden:
        name = name.replace(ch, "")
    return name[:31]


def _to_excel_value(value):
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (dict, list, tuple)):
        try:
            return json.dumps(value, ensure_ascii=False, default=str)
        except (TypeError, ValueError):
            return str(value)
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return str(value)
