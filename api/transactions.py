"""API роуты для транзакций"""

from fastapi import APIRouter, Request, Query, UploadFile, File, Form
from fastapi.responses import (
    HTMLResponse,
    RedirectResponse,
    JSONResponse,
    StreamingResponse,
)
from fastapi.templating import Jinja2Templates
from main import get_db, get_user
from typing import Optional
import logging
import openpyxl
import io
import uuid
import os
from datetime import datetime

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/transactions", tags=["transactions"])
templates = Jinja2Templates(directory="templates")

# Директория для загрузки
UPLOAD_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "uploads"
)
os.makedirs(UPLOAD_DIR, exist_ok=True)


@router.get("/", response_class=HTMLResponse)
async def transactions_list(
    request: Request, filter_type: Optional[str] = Query("all"), page: int = Query(1)
):
    user = get_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    db = get_db()
    try:
        transactions = db.get_transactions_dict(limit=1000)
    except Exception as e:
        logger.error(f"Transactions load error: {e}")
        transactions = []

    if filter_type != "all":
        transactions = [
            t for t in transactions if t.get("operation_type") == filter_type
        ]

    per_page = 50
    total = len(transactions)
    start = (page - 1) * per_page
    paginated = transactions[start : start + per_page]

    return templates.TemplateResponse(
        "transactions/list.html",
        {
            "request": request,
            "current_user": user,
            "transactions": paginated,
            "filter_type": filter_type,
            "page": page,
            "total_pages": max(1, (total + per_page - 1) // per_page),
            "total_transactions": total,
        },
    )


@router.get("/export")
async def export_operations(request: Request, export_type: str = Query("all")):
    """Экспорт операций в Excel (all/income/expense)"""
    user = get_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    db = get_db()
    try:
        transactions = db.get_transactions_dict(limit=10000)

        if export_type == "income":
            transactions = [
                t for t in transactions if t.get("operation_type") == "income"
            ]
            filename_prefix = "income"
        elif export_type == "expense":
            transactions = [
                t for t in transactions if t.get("operation_type") == "expense"
            ]
            filename_prefix = "expense"
        else:
            filename_prefix = "all_operations"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Операции"

        # Заголовки
        ws.append(
            [
                "Дата",
                "Товар ID",
                "Товар",
                "Тип",
                "Количество",
                "Пользователь",
                "Примечание",
            ]
        )

        for txn in transactions:
            ws.append(
                [
                    txn.get("timestamp", ""),
                    txn.get("item_id", ""),
                    txn.get("item_name", ""),
                    txn.get("operation_type", ""),
                    txn.get("quantity", 0),
                    txn.get("user_name", ""),
                    txn.get("detail", ""),
                ]
            )

        # Formatting
        for col_idx in range(1, 8):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

        ws.auto_filter.ref = f"A1:G{len(transactions) + 1}"
        ws.freeze_panes = "A2"

        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
        filename = f"operations_{filename_prefix}_{timestamp}.xlsx"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        logger.error(f"Export operations error: {e}")
        return HTMLResponse(content=f"Ошибка экспорта: {str(e)}", status_code=500)


@router.get("/import")
async def import_operations_page(request: Request):
    """Страница импорта операций из Excel (только для администратора)"""
    user = get_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if user.get("role") != "admin":
        return RedirectResponse(url="/transactions", status_code=303)
    return templates.TemplateResponse(
        "transactions/import.html", {"request": request, "current_user": user}
    )


@router.post("/import/upload")
async def import_operations_upload(
    request: Request, file: UploadFile = File(...), operation_type: str = Form("income")
):
    """Загрузка и парсинг Excel файла операций (только для администратора)"""
    user = get_user(request)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=403)
    if user.get("role") != "admin":
        return JSONResponse(
            {"error": "Только администратор может импортировать операции"},
            status_code=403,
        )

    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        return JSONResponse(
            {"error": "Поддерживаются только .xlsx и .xls файлы"}, status_code=400
        )

    try:
        content = await file.read()
        if len(content) > 10 * 1024 * 1024:
            return JSONResponse(
                {"error": "Файл слишком большой (макс. 10MB)"}, status_code=400
            )

        # Сохраняем временный файл
        file_id = str(uuid.uuid4())
        file_ext = os.path.splitext(file.filename)[1]
        temp_path = os.path.join(UPLOAD_DIR, f"{file_id}{file_ext}")

        with open(temp_path, "wb") as f:
            f.write(content)

        # Парсим Excel
        wb = openpyxl.load_workbook(temp_path, data_only=True)
        ws = wb.active

        # Читаем заголовки
        headers = [cell.value for cell in ws[1]]

        # Читаем данные
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                row_dict = {headers[i]: row[i] for i in range(len(headers))}
                row_dict["operation_type"] = operation_type
                rows.append(row_dict)

        wb.close()

        # Сохраняем в session
        request.session[f"import_{file_id}"] = {
            "file_path": temp_path,
            "operation_type": operation_type,
            "headers": headers,
            "row_count": len(rows),
        }

        return JSONResponse(
            {
                "file_id": file_id,
                "headers": headers,
                "row_count": len(rows),
                "preview": rows[:10],
                "operation_type": operation_type,
            }
        )

    except Exception as e:
        logger.error(f"Import upload error: {e}")
        return JSONResponse({"error": f"Ошибка загрузки: {str(e)}"}, status_code=500)


@router.post("/import/execute")
async def import_operations_execute(request: Request, file_id: str = Form(...)):
    """Выполнение импорта операций (только для администратора)"""
    user = get_user(request)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=403)
    if user.get("role") != "admin":
        return JSONResponse(
            {"error": "Только администратор может импортировать операции"},
            status_code=403,
        )

    import_data = request.session.get(f"import_{file_id}")
    if not import_data:
        return JSONResponse({"error": "Данные импорта не найдены"}, status_code=404)

    db = get_db()
    try:
        rows = import_data.get("row_count", 0)
        operation_type = import_data.get("operation_type", "income")
        success_count = 0
        error_count = 0

        # Открываем файл снова
        wb = openpyxl.load_workbook(import_data["file_path"], data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                if not any(cell is not None for cell in row):
                    continue

                row_dict = {headers[i]: row[i] for i in range(len(headers))}

                # Маппинг колонок (гибкий)
                item_id = (
                    row_dict.get("item_id")
                    or row_dict.get("ID товара")
                    or row_dict.get("item_code")
                )
                quantity = int(
                    row_dict.get("quantity")
                    or row_dict.get("Количество")
                    or row_dict.get("qty")
                    or 0
                )
                detail = (
                    row_dict.get("detail")
                    or row_dict.get("Примечание")
                    or row_dict.get("note")
                    or ""
                )
                user_name = (
                    row_dict.get("user")
                    or row_dict.get("Пользователь")
                    or user.get("username", "")
                )

                if not item_id or quantity <= 0:
                    error_count += 1
                    continue

                # Выполняем операцию
                if operation_type == "income":
                    db.income_item(item_id, quantity, user.get("id"), detail)
                else:
                    db.expense_item(item_id, quantity, user.get("id"), detail)

                success_count += 1
            except Exception as e:
                logger.error(f"Import row {row_idx} error: {e}")
                error_count += 1

        wb.close()

        # Удаляем временный файл
        try:
            os.remove(import_data["file_path"])
        except:
            pass

        # Удаляем из session
        request.session.pop(f"import_{file_id}", None)

        return JSONResponse(
            {
                "success": True,
                "success_count": success_count,
                "error_count": error_count,
                "operation_type": operation_type,
            }
        )

    except Exception as e:
        logger.error(f"Import execute error: {e}")
        return JSONResponse({"error": f"Ошибка импорта: {str(e)}"}, status_code=500)


@router.delete("/api/transactions/{transaction_id}")
async def delete_transaction(request: Request, transaction_id: int):
    """Удаление транзакции (только для администратора)"""
    user = get_user(request)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    if user.get("role") != "admin":
        return JSONResponse(
            {"error": "Только администратор может удалять операции"}, status_code=403
        )

    db = get_db()
    try:
        success = db.delete_transaction(transaction_id)
        if success:
            return JSONResponse({"success": True})
        else:
            return JSONResponse({"error": "Операция не найдена"}, status_code=404)
    except Exception as e:
        logger.error(f"Delete transaction error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)
