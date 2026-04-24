"""
Сервис импорта данных из Excel файлов
"""
import logging
import openpyxl
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
import json

logger = logging.getLogger(__name__)


class ExcelImportService:
    """Сервис для импорта товаров из Excel файлов"""
    
    # Маппинг возможных названий колонок в Excel на поля Item
    COLUMN_MAPPINGS = {
        'item_id': ['item_id', 'код', 'артикул', 'id', 'идентификатор'],
        'name': ['name', 'название', 'наименование', 'товар'],
        'quantity': ['quantity', 'количество', 'остаток', 'кол-во'],
        'min_stock': ['min_stock', 'минимум', 'мин_остаток', 'минимальный'],
        'category': ['category', 'категория', 'группа', 'тип'],
        'location': ['location', 'расположение', 'место', 'склад'],
        'image_url': ['image_url', 'изображение', 'фото', 'картинка'],
        'shop_url': ['shop_url', 'ссылка', 'магазин', 'url'],
        'specifications': ['specifications', 'характеристики', 'specs', 'параметры'],
    }
    
    def __init__(self):
        self.parsed_data: List[Dict[str, Any]] = []
        self.headers: List[str] = []
        self.column_mapping: Dict[str, str] = {}
        self.errors: List[Dict[str, Any]] = []
        self.stats = {
            'total_rows': 0,
            'valid_rows': 0,
            'invalid_rows': 0,
            'duplicate_rows': 0,
        }
    
    def parse_excel_file(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        handle_merged_cells: bool = True
    ) -> Dict[str, Any]:
        """
        Парсинг Excel файла
        
        Returns:
        {
            'sheets': [имена листов],
            'headers': [заголовки],
            'data': [данные с заголовками],
            'row_count': количество строк
        }
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Определяем нужный лист
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            logger.info(f"Parsing sheet: {ws.title}")
            
            # Читаем все строки с обработкой объединенных ячеек
            raw_data = []
            if handle_merged_cells:
                raw_data = self._read_with_merged_cells(ws)
            else:
                raw_data = list(ws.iter_rows(values_only=True))
            
            if not raw_data or len(raw_data) < 2:
                return {
                    'sheets': wb.sheetnames,
                    'headers': [],
                    'data': [],
                    'row_count': 0,
                    'error': 'Файл пуст или не содержит данных'
                }
            
            # Первая строка - заголовки
            self.headers = [str(h).strip() if h else '' for h in raw_data[0]]
            
            # Остальные строки - данные
            data_rows = raw_data[1:]
            
            # Преобразуем в словари
            parsed = []
            for row_idx, row in enumerate(data_rows, start=2):
                row_dict = {}
                for col_idx, header in enumerate(self.headers):
                    if col_idx < len(row):
                        row_dict[header] = row[col_idx]
                    else:
                        row_dict[header] = None
                row_dict['_row_num'] = row_idx
                parsed.append(row_dict)
            
            self.parsed_data = parsed
            
            return {
                'sheets': wb.sheetnames,
                'headers': self.headers,
                'data': parsed,
                'row_count': len(parsed)
            }
            
        except Exception as e:
            logger.error(f"Excel parse error: {e}")
            return {
                'sheets': [],
                'headers': [],
                'data': [],
                'row_count': 0,
                'error': f'Ошибка чтения файла: {str(e)}'
            }
    
    def _read_with_merged_cells(self, ws) -> List[List[Any]]:
        """Чтение данных с учетом объединенных ячеек (forward fill)"""
        data = []
        
        # Заполняем объединенные ячейки
        merged_values = {}
        for merged_range in ws.merged_cells.ranges:
            value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_values[(row, col)] = value
        
        # Читаем все строки
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            row_data = []
            for cell in row:
                if (cell.row, cell.column) in merged_values:
                    row_data.append(merged_values[(cell.row, cell.column)])
                else:
                    row_data.append(cell.value)
            data.append(row_data)
        
        return data
    
    def auto_detect_mapping(self) -> Dict[str, str]:
        """Автоматическое определение маппинга колонок"""
        mapping = {}
        
        for field, possible_names in self.COLUMN_MAPPINGS.items():
            for header in self.headers:
                header_lower = header.lower().strip()
                if header_lower in possible_names:
                    mapping[field] = header
                    break
        
        self.column_mapping = mapping
        return mapping
    
    def validate_and_map_data(
        self,
        column_mapping: Optional[Dict[str, str]] = None,
        existing_item_ids: Optional[set] = None
    ) -> Dict[str, Any]:
        """
        Валидация и маппинг данных
        
        Args:
            column_mapping: {field_name: excel_header}
            existing_item_ids: множество существующих item_id для проверки дубликатов
        
        Returns:
        {
            'valid_items': [словари товаров],
            'errors': [ошибки валидации],
            'stats': статистика
        }
        """
        if column_mapping:
            self.column_mapping = column_mapping
        
        self.errors = []
        valid_items = []
        
        self.stats = {
            'total_rows': len(self.parsed_data),
            'valid_rows': 0,
            'invalid_rows': 0,
            'duplicate_rows': 0,
        }
        
        for row_data in self.parsed_data:
            row_num = row_data.get('_row_num', 0)
            
            try:
                # Маппинг полей
                item = {}
                for field, header in self.column_mapping.items():
                    value = row_data.get(header)
                    if value is not None:
                        item[field] = self._clean_value(value)
                
                # Валидация обязательных полей
                if not item.get('item_id'):
                    self.errors.append({
                        'row': row_num,
                        'error': 'Отсутствует item_id (код/артикул)',
                        'data': row_data
                    })
                    self.stats['invalid_rows'] += 1
                    continue
                
                if not item.get('name'):
                    self.errors.append({
                        'row': row_num,
                        'error': 'Отсутствует название',
                        'data': row_data
                    })
                    self.stats['invalid_rows'] += 1
                    continue
                
                # Проверка на дубликаты
                if existing_item_ids and item['item_id'] in existing_item_ids:
                    self.stats['duplicate_rows'] += 1
                    item['is_duplicate'] = True
                
                # Преобразование типов
                if 'quantity' in item:
                    try:
                        item['quantity'] = int(item['quantity'])
                    except (ValueError, TypeError):
                        item['quantity'] = 0
                
                if 'min_stock' in item:
                    try:
                        item['min_stock'] = int(item['min_stock'])
                    except (ValueError, TypeError):
                        item['min_stock'] = 1
                
                # Specifications как JSON
                if 'specifications' in item and isinstance(item['specifications'], dict):
                    item['specifications'] = json.dumps(item['specifications'], ensure_ascii=False)
                
                item['is_valid'] = True
                valid_items.append(item)
                self.stats['valid_rows'] += 1
                
            except Exception as e:
                self.errors.append({
                    'row': row_num,
                    'error': f'Ошибка обработки: {str(e)}',
                    'data': row_data
                })
                self.stats['invalid_rows'] += 1
        
        return {
            'valid_items': valid_items,
            'errors': self.errors,
            'stats': self.stats,
            'column_mapping': self.column_mapping
        }
    
    def _clean_value(self, value: Any) -> Any:
        """Очистка значения"""
        if value is None:
            return None
        if isinstance(value, str):
            value = value.strip()
            return value if value else None
        if isinstance(value, datetime):
            return value.isoformat()
        return value
    
    def execute_import(
        self,
        db,
        items: List[Dict[str, Any]],
        update_duplicates: bool = False,
        batch_size: int = 100
    ) -> Dict[str, Any]:
        """
        Выполнение импорта в БД
        
        Args:
            db: DatabaseManager
            items: список валидных товаров
            update_duplicates: обновлять ли существующие товары
            batch_size: размер пакета для коммита
        
        Returns:
        {
            'created': количество созданных,
            'updated': количество обновленных,
            'skipped': количество пропущенных,
            'errors': список ошибок
        }
        """
        result = {
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': [],
            'total': len(items)
        }
        
        try:
            # Инвалидируем кэш товаров
            db.invalidate_items_cache()
            
            for idx, item in enumerate(items):
                try:
                    item_id = item['item_id']
                    existing = db.get_item_by_id(item_id)
                    
                    if existing:
                        if update_duplicates:
                            # Обновляем существующий товар
                            success = db.update_item(
                                item_id=item_id,
                                name=item.get('name'),
                                quantity=item.get('quantity'),
                                min_stock=item.get('min_stock'),
                                category=item.get('category'),
                                location=item.get('location'),
                                image_url=item.get('image_url'),
                                shop_url=item.get('shop_url'),
                                specifications=item.get('specifications')
                            )
                            if success:
                                result['updated'] += 1
                            else:
                                result['errors'].append({
                                    'item_id': item_id,
                                    'error': 'Ошибка обновления'
                                })
                        else:
                            result['skipped'] += 1
                    else:
                        # Создаем новый товар
                        new_item = db.create_item(
                            item_id=item_id,
                            name=item.get('name', ''),
                            quantity=item.get('quantity', 0),
                            min_stock=item.get('min_stock', 1),
                            category=item.get('category'),
                            location=item.get('location')
                        )
                        
                        if new_item:
                            # Обновляем дополнительные поля
                            if item.get('image_url'):
                                db.update_item_field(item_id, 'image_url', item['image_url'])
                            if item.get('shop_url'):
                                db.update_item_field(item_id, 'shop_url', item['shop_url'])
                            if item.get('specifications'):
                                db.update_item_field(item_id, 'specifications', item['specifications'])
                            
                            result['created'] += 1
                        else:
                            result['errors'].append({
                                'item_id': item_id,
                                'error': 'Ошибка создания'
                            })
                    
                    # Периодический коммит и инвалидация кэша
                    if (idx + 1) % batch_size == 0:
                        db.invalidate_items_cache()
                        
                except Exception as e:
                    logger.error(f"Import error for item {item.get('item_id')}: {e}")
                    result['errors'].append({
                        'item_id': item.get('item_id', 'unknown'),
                        'error': str(e)
                    })
            
            # Финальная инвалидация кэша
            db.invalidate_items_cache()
            
            logger.info(f"Import completed: {result['created']} created, {result['updated']} updated, {result['skipped']} skipped")
            
        except Exception as e:
            logger.error(f"Import execution error: {e}")
            result['errors'].append({
                'item_id': 'general',
                'error': f'Критическая ошибка импорта: {str(e)}'
            })
        
        return result
