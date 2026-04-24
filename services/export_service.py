"""
Сервис экспорта данных БД в Excel (.xlsx)
"""

import os
import logging
from datetime import datetime
from typing import Dict, List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

EXPORTS_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "..", "exports"
)
EXPORTS_DIR = os.path.normpath(EXPORTS_DIR)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
ALT_ROW_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


def export_all_tables_to_xlsx(tables_data: Dict[str, List[Dict]]) -> str:
    """
    Экспортирует все таблицы БД в один .xlsx файл.
    Каждая таблица — отдельный лист.

    Args:
        tables_data: {table_name: [list_of_row_dicts]}

    Returns:
        str: Путь к созданному файлу
    """
    os.makedirs(EXPORTS_DIR, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    filename = f"sklad_export_{timestamp}.xlsx"
    filepath = os.path.join(EXPORTS_DIR, filename)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for table_name, rows in tables_data.items():
        ws = wb.create_sheet(title=_sanitize_sheet_name(table_name))

        if not rows:
            ws.cell(row=1, column=1, value="Нет данных")
            ws.cell(row=1, column=1).font = Font(italic=True, color="999999")
            continue

        columns = list(rows[0].keys())

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row_data.get(col_name, "")
                value = _to_excel_value(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = DATA_ALIGNMENT
                cell.border = THIN_BORDER
                if row_idx % 2 == 0:
                    cell.fill = ALT_ROW_FILL

        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

        ws.auto_filter.ref = (
            f"A1:{openpyxl.utils.get_column_letter(len(columns))}{len(rows) + 1}"
        )
        ws.freeze_panes = "A2"

    wb.save(filepath)
    logger.info(f"Exported {len(tables_data)} tables to {filepath}")
    return filepath


def _sanitize_sheet_name(name: str) -> str:
    """
    Приводит имя таблицы к допустимому имени листа Excel.
    Максимум 31 символ, без запрещённых символов.
    """
    forbidden = [":", "\\", "/", "?", "*", "[", "]"]
    for ch in forbidden:
        name = name.replace(ch, "")
    return name[:31]


def _to_excel_value(value):
    """
    Преобразует значение в формат, совместимый с openpyxl.
    """
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (dict, list, tuple)):
        import json

        try:
            return json.dumps(value, ensure_ascii=False, default=str)
        except (TypeError, ValueError):
            return str(value)
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return str(value)
